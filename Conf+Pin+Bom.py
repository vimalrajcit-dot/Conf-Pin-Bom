import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import io


st.set_page_config(page_title="BOM + Configuration Processor", layout="wide")

st.title("📦 BOM + Configuration Processor")
st.write("Upload BOM and Configuration files. Process data. Download result.")


# -------------------------------------------------------
# HELPER FUNCTIONS
# -------------------------------------------------------

def process_files(bom_df, config_df):

    # Identify columns
    if 'Tag' not in config_df.columns:
        tag_cols = [c for c in config_df.columns if c.lower() == 'tag']
        if tag_cols:
            config_df.rename(columns={tag_cols[0]: 'Tag'}, inplace=True)
        else:
            st.error("No Tag column found in Configuration file")
            return None

    if 'Part Name' not in bom_df.columns:
        st.error("No 'Part Name' column in BOM file")
        return None

    if 'Part Number' not in bom_df.columns:
        st.error("No 'Part Number' column in BOM file")
        return None

    # Tag column in BOM
    tag_col = 'Tag Number' if 'Tag Number' in bom_df.columns else None
    if tag_col is None:
        for c in bom_df.columns:
            if c.lower() == 'tag number':
                tag_col = c
                break

    if tag_col is None:
        st.error("No Tag Number column in BOM file")
        return None

    # Build mapping
    tag_mapping = {}
    for _, row in bom_df.iterrows():
        tag = str(row[tag_col]).strip()
        part = row['Part Name']
        number = row['Part Number']

        if pd.notna(tag) and pd.notna(part) and pd.notna(number):
            if tag not in tag_mapping:
                tag_mapping[tag] = {}
            tag_mapping[tag][part] = number

    part_names = bom_df['Part Name'].dropna().unique()

    # Add columns
    for part in part_names:
        if part not in config_df.columns:
            config_df[part] = None

    # Fill data
    progress = st.progress(0)
    total = len(config_df)
    filled = 0

    for i, row in config_df.iterrows():
        tag = str(row['Tag']).strip()

        if tag in tag_mapping:
            for part, number in tag_mapping[tag].items():
                config_df.at[i, part] = number
                filled += 1

        progress.progress((i + 1) / total)

    st.success(f"Filled {filled} cells")

    return config_df, part_names


def format_excel(output_file, part_names):

    wb = load_workbook(output_file)
    ws = wb.active

    yellow_fill = PatternFill(start_color="FFFF00",
                              end_color="FFFF00",
                              fill_type="solid")

    header_map = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}

    for part in part_names:
        col = header_map.get(part)
        if col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value in [None, ""]:
                    cell.fill = yellow_fill

    # Auto adjust columns
    for col in range(1, ws.max_column + 1):
        max_length = 0
        letter = get_column_letter(col)

        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[letter].width = min(max_length + 2, 50)

    wb.save(output_file)


# -------------------------------------------------------
# STREAMLIT UI
# -------------------------------------------------------

uploaded_bom = st.file_uploader("Upload BOM Excel", type=["xlsx"])
uploaded_config = st.file_uploader("Upload Configuration Excel", type=["xlsx"])

if uploaded_bom and uploaded_config:

    bom_df = pd.read_excel(uploaded_bom)
    config_df = pd.read_excel(uploaded_config)

    if st.button("🚀 Run Process"):

        result = process_files(bom_df, config_df)

        if result:

            processed_df, part_names = result

            output = io.BytesIO()
            processed_df.to_excel(output, index=False)
            output.seek(0)

            # Save temp file for formatting
            temp_path = "output_temp.xlsx"
            processed_df.to_excel(temp_path, index=False)

            format_excel(temp_path, part_names)

            with open(temp_path, "rb") as f:
                data = f.read()

            st.download_button(
                label="⬇️ Download Result",
                data=data,
                file_name="BOM_Configuration_Result.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            st.success("Processing completed!")